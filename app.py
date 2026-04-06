import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import calendar
import re
import math

# ─────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Excelhub HR Automation",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────
LUNCH_HRS  = 0.5   # deducted from every shift
STD_SHIFT  = 8.0   # hours after lunch → OT starts
MAX_OT     = 3.3   # max OT hours per day
WO_MULT    = 1.5   # weekoff / holiday multiplier
TOTAL_SHIFT = 12.0  # total shift hours

# ─────────────────────────────────────────────────────────────
#  CUSTOM STYLES
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stAppViewContainer"] { background: #f7f8fa; }
    [data-testid="stHeader"] { background: transparent; }
    .main .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

    .app-header {
        background: linear-gradient(135deg, #1a2942 0%, #243859 100%);
        color: white;
        padding: 20px 28px;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        display: flex;
        align-items: center;
        gap: 16px;
    }
    .app-header h1 { margin: 0; font-size: 22px; font-weight: 600; }
    .app-header p  { margin: 4px 0 0; font-size: 13px; opacity: 0.75; }

    .step-bar {
        display: flex;
        gap: 0;
        margin-bottom: 1.5rem;
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid #e0e4ec;
    }
    .step-item {
        flex: 1;
        padding: 10px 14px;
        font-size: 13px;
        font-weight: 500;
        text-align: center;
        background: white;
        color: #888;
        border-right: 1px solid #e0e4ec;
        cursor: default;
    }
    .step-item:last-child { border-right: none; }
    .step-item.done { background: #e8f5e9; color: #2e7d32; }
    .step-item.active { background: #1a2942; color: white; }

    .info-card {
        background: white;
        border: 1px solid #e0e4ec;
        border-radius: 10px;
        padding: 16px 18px;
        margin-bottom: 12px;
    }
    .info-card h4 { margin: 0 0 4px; font-size: 14px; color: #1a2942; }
    .info-card p  { margin: 0; font-size: 12px; color: #666; line-height: 1.5; }

    .badge-ot   { background:#e8f5e9; color:#2e7d32; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
    .badge-wot  { background:#fff3e0; color:#e65100; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
    .badge-miss { background:#fff9c4; color:#f57f17; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
    .badge-abs  { background:#fce4ec; color:#c62828; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }

    .rule-box {
        background: #f0f4ff;
        border-left: 3px solid #1a2942;
        padding: 10px 14px;
        border-radius: 0 8px 8px 0;
        font-size: 13px;
        color: #333;
        margin: 8px 0;
        font-family: monospace;
        line-height: 1.8;
    }

    .anomaly-row {
        background: #fff8e1;
        border-left: 3px solid #ffa000;
        padding: 8px 12px;
        border-radius: 0 6px 6px 0;
        font-size: 13px;
        margin: 4px 0;
    }

    div[data-testid="stButton"] > button {
        border-radius: 8px;
        font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
#  SESSION STATE INIT
# ─────────────────────────────────────────────────────────────
defaults = {
    'step': 1,
    'parsed': {},
    'emp_map': {},
    'holidays': [],
    'month': datetime.today().month,
    'year':  datetime.today().year,
    'attendance': None,
    'corrections': {},
    'basic_details': None,
    'att_bytes': None,
    'sal_bytes': None,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─────────────────────────────────────────────────────────────
#  BUSINESS LOGIC FUNCTIONS
# ─────────────────────────────────────────────────────────────

def parse_dat(content: str) -> dict:
    """
    Parse biometric .dat file.
    Returns: {emp_id_str: {date_str: [datetime, ...]}}
    Handles tab / comma / pipe / multi-space delimited formats.
    """
    records = defaultdict(lambda: defaultdict(list))
    for line in content.strip().splitlines():
        line = line.strip()
        if not line or line.startswith('#') or line.lower().startswith('no'):
            continue
        parts = re.split(r'[\t,|]+|\s{2,}', line)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) < 3:
            continue
        try:
            raw_id = re.sub(r'\D', '', parts[0])
            if not raw_id:
                continue
            emp_id = str(int(raw_id)).zfill(4)

            raw_date = parts[1]
            if len(raw_date) == 8 and raw_date.isdigit():
                d = datetime.strptime(raw_date, '%Y%m%d').date()
            elif '-' in raw_date:
                d = datetime.strptime(raw_date[:10], '%Y-%m-%d').date()
            elif '/' in raw_date:
                fmt = '%d/%m/%Y' if len(raw_date.split('/')[0]) == 2 else '%Y/%m/%d'
                d = datetime.strptime(raw_date, fmt).date()
            else:
                continue

            raw_time = re.sub(r'[^\d]', '', parts[2])
            if len(raw_time) == 4:
                raw_time += '00'
            t = datetime.strptime(raw_time[:6], '%H%M%S').time()

            records[emp_id][str(d)].append(datetime.combine(d, t))
        except Exception:
            continue
    return dict(records)


def calc_ot(in_dt, out_dt, is_weekoff: bool, ot_rate: float):
    """
    Shift logic:
      Total shift = 12 hrs
      Standard work = 8 hrs + 0.5 lunch
      OT = max(0, net_hrs - 8), capped at 3.3
      Weekoff OT rate = 1.5x regular
    Returns: (ot_hrs, ot_pay, ot_type, net_hrs)
    """
    if isinstance(in_dt, str):
        in_dt = datetime.strptime(in_dt, '%H:%M')
    if isinstance(out_dt, str):
        out_dt = datetime.strptime(out_dt, '%H:%M')

    if out_dt <= in_dt:
        out_dt += timedelta(hours=24)

    gross_hrs = (out_dt - in_dt).total_seconds() / 3600
    net_hrs   = max(0, gross_hrs - LUNCH_HRS)
    raw_ot    = max(0, net_hrs - STD_SHIFT)
    ot_hrs    = min(raw_ot, MAX_OT)

    rate    = ot_rate * WO_MULT if is_weekoff else ot_rate
    ot_pay  = round(ot_hrs * rate, 2)
    ot_type = 'weekoff' if is_weekoff else ('regular' if ot_hrs > 0 else None)

    return round(ot_hrs, 2), ot_pay, ot_type, round(net_hrs, 2)


def read_emp_map(att_bytes: bytes) -> dict:
    """
    Read Sheet3 from attendance Excel.
    Returns: {bio_id: {name, unit, sl, ot_rate}}
    Sheet3 layout has UNIT1 cols (1,2,3) and UNIT2 cols (5,6,7) side by side.
    """
    wb  = openpyxl.load_workbook(io.BytesIO(att_bytes), data_only=True)
    emp = {}
    if 'Sheet3' not in wb.sheetnames:
        return emp

    ws = wb['Sheet3']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Unit-1: col B(sl), C(name), D(bio_id) → indices 1,2,3
        try:
            if row[2] and row[3]:
                bio = str(int(float(str(row[3])))).zfill(4)
                emp[bio] = {'name': str(row[2]).strip(), 'unit': 'Unit-1',
                            'sl': row[1], 'ot_rate': 0}
        except Exception:
            pass
        # Unit-2: col F(sl), G(name), H(bio_id) → indices 5,6,7
        try:
            if len(row) > 7 and row[6] and row[7]:
                bio = str(int(float(str(row[7])))).zfill(4)
                emp[bio] = {'name': str(row[6]).strip(), 'unit': 'Unit-2',
                            'sl': row[5], 'ot_rate': 0}
        except Exception:
            pass
    return emp


def load_ot_rates(sal_bytes: bytes, emp_map: dict) -> dict:
    """Merge OT rates from Basic Details sheet into emp_map."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(sal_bytes), data_only=True)
        if 'Basic Details' not in wb.sheetnames:
            return emp_map

        ws     = wb['Basic Details']
        rows   = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c else '' for c in rows[0]]

        try:
            name_col = header.index('Employee Name')
            ot_col   = header.index('OT Per HR')
        except ValueError:
            return emp_map

        name_to_rate = {}
        for row in rows[1:]:
            if row[0] is None:
                continue
            name = str(row[name_col]).strip().lower() if row[name_col] else ''
            rate = row[ot_col]
            try:
                rate = float(rate) if rate else 0
            except Exception:
                rate = 0
            if name:
                name_to_rate[name] = rate

        for bio_id, info in emp_map.items():
            match = info['name'].strip().lower()
            if match in name_to_rate:
                emp_map[bio_id]['ot_rate'] = name_to_rate[match]
    except Exception:
        pass
    return emp_map


def build_attendance(parsed: dict, emp_map: dict, year: int, month: int,
                     holidays: list, corrections: dict):
    """
    Build list of employee attendance dicts.
    Each dict has daily records + summary.
    """
    num_days = calendar.monthrange(year, month)[1]
    dates    = [date(year, month, d) for d in range(1, num_days + 1)]
    holiday_set = set(holidays)
    rows = []

    for bio_id, info in sorted(emp_map.items(),
                                key=lambda x: (x[1].get('unit', ''), x[1].get('sl') or 999)):
        ot_rate   = info.get('ot_rate', 0) or 0
        emp_dates = parsed.get(bio_id, {})
        day_records = []

        for d in dates:
            is_sunday  = d.weekday() == 6
            is_holiday = d in holiday_set
            is_weekoff = is_sunday or is_holiday
            d_str      = str(d)

            corr_key = (bio_id, d_str)
            if corr_key in corrections:
                in_dt, out_dt = corrections[corr_key]
                status = 'corrected'
            elif d_str in emp_dates:
                punches = sorted(emp_dates[d_str])
                # Remove duplicate punches within 5 minutes
                deduped = [punches[0]]
                for p in punches[1:]:
                    if (p - deduped[-1]).total_seconds() > 300:
                        deduped.append(p)
                in_dt  = deduped[0]
                out_dt = deduped[-1] if len(deduped) > 1 else None
                if in_dt and out_dt and in_dt != out_dt:
                    status = 'ok'
                elif in_dt and not out_dt:
                    status = 'missing_out'
                else:
                    status = 'missing_in'
            else:
                in_dt = out_dt = None
                if is_sunday:
                    status = 'sunday'
                elif is_holiday:
                    status = 'holiday'
                else:
                    status = 'absent'

            # OT calc
            ot_hrs = ot_pay = net_hrs = 0
            ot_type = None
            if in_dt and out_dt and in_dt != out_dt and status not in ('absent',):
                ot_hrs, ot_pay, ot_type, net_hrs = calc_ot(in_dt, out_dt, is_weekoff, ot_rate)

            def fmt_time(dt):
                if dt is None:
                    return None
                if isinstance(dt, datetime):
                    return dt.strftime('%H:%M')
                if isinstance(dt, str):
                    return dt[:5]
                return str(dt)[:5]

            day_records.append({
                'date':       d,
                'day':        d.strftime('%a'),
                'in_time':    fmt_time(in_dt),
                'out_time':   fmt_time(out_dt),
                'status':     status,
                'is_weekoff': is_weekoff,
                'ot_hrs':     ot_hrs,
                'ot_pay':     ot_pay,
                'ot_type':    ot_type,
                'net_hrs':    net_hrs,
            })

        present   = sum(1 for r in day_records if r['status'] in ('ok', 'corrected'))
        absent    = sum(1 for r in day_records if r['status'] == 'absent')
        reg_ot    = round(sum(r['ot_hrs'] for r in day_records if r['ot_type'] == 'regular'), 2)
        wkoff_ot  = round(sum(r['ot_hrs'] for r in day_records if r['ot_type'] == 'weekoff'), 2)
        mp_days   = [r for r in day_records if 'missing' in r['status']]

        rows.append({
            'bio_id':          bio_id,
            'name':            info['name'],
            'unit':            info['unit'],
            'ot_rate':         ot_rate,
            'day_records':     day_records,
            'present':         present,
            'absent':          absent,
            'reg_ot':          reg_ot,
            'wkoff_ot':        wkoff_ot,
            'misspunch_count': len(mp_days),
            'misspunch_days':  mp_days,
        })
    return rows


# ─────────────────────────────────────────────────────────────
#  EXCEL GENERATORS
# ─────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def generate_attendance_excel(emp_rows: list, year: int, month: int) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    num_days  = calendar.monthrange(year, month)[1]
    dates     = [date(year, month, d) for d in range(1, num_days + 1)]
    month_lbl = datetime(year, month, 1).strftime('%b %Y')

    HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=9)
    SUN_FILL  = PatternFill('solid', fgColor='EDE7F6')
    HOL_FILL  = PatternFill('solid', fgColor='E8EAF6')
    ABS_FILL  = PatternFill('solid', fgColor='FFEBEE')
    IN_FILL   = PatternFill('solid', fgColor='E8F5E9')
    OUT_FILL  = PatternFill('solid', fgColor='FFF3E0')
    OT_FILL   = PatternFill('solid', fgColor='E3F2FD')
    WOT_FILL  = PatternFill('solid', fgColor='FFF9C4')
    MP_FILL   = PatternFill('solid', fgColor='FFF9C4')
    border    = _thin_border()
    center    = _center()

    for unit in ('Unit-1', 'Unit-2'):
        unit_emps = [e for e in emp_rows if e['unit'] == unit]
        if not unit_emps:
            continue

        ws = wb.create_sheet(unit)
        ws.freeze_panes = 'E3'

        # Row 1: month label
        ws.cell(1, 1, month_lbl).font = Font(bold=True, size=12, color='1F4E79')

        # Row 2: headers
        day_headers = [f"{d.strftime('%a')}\n{d.day}" for d in dates]
        all_headers = ['SL', 'Employee Name', 'Designation', ''] + day_headers + \
                      ['Reg OT', 'Wkoff OT', 'Present', 'Absent']

        for col, h in enumerate(all_headers, 1):
            c = ws.cell(2, col, h)
            c.fill   = HDR_FILL
            c.font   = HDR_FONT
            c.alignment = center
            c.border = border
            # highlight Sundays in header
            if col >= 5 and col <= 4 + num_days:
                d = dates[col - 5]
                if d.weekday() == 6:
                    c.fill = PatternFill('solid', fgColor='4A148C')

        ws.row_dimensions[2].height = 28
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 5
        for col in range(5, 5 + num_days):
            ws.column_dimensions[get_column_letter(col)].width = 6.5
        for i, lbl in enumerate(['Reg OT', 'Wkoff OT', 'Present', 'Absent']):
            ws.column_dimensions[get_column_letter(5 + num_days + i)].width = 10

        row_num = 3
        for sl_num, emp in enumerate(unit_emps, 1):
            # Merge SL + Name across 3 rows visually
            for r in (row_num, row_num+1, row_num+2):
                for c in (1, 2, 3):
                    cell = ws.cell(r, c)
                    cell.border = border
                    cell.alignment = center
                    cell.font = Font(size=9)

            ws.cell(row_num, 1, sl_num).font = Font(size=9, bold=True)
            ws.cell(row_num, 2, emp['name']).font = Font(size=9, bold=True)

            # IN / OUT / OT labels
            for r, lbl in zip((row_num, row_num+1, row_num+2), ('IN', 'OUT', 'OT')):
                c = ws.cell(r, 4, lbl)
                c.font      = Font(size=8, bold=True, color='1F4E79')
                c.alignment = center
                c.border    = border

            reg_ot_total = wkoff_ot_total = present_count = absent_count = 0

            for col_offset, dr in enumerate(emp['day_records']):
                col = 5 + col_offset
                st_  = dr['status']
                is_wo = dr['is_weekoff']

                # Determine cell values
                if st_ == 'sunday':
                    in_v = out_v = 'S'
                    fi = fo = SUN_FILL
                elif st_ == 'holiday':
                    in_v = out_v = 'H'
                    fi = fo = HOL_FILL
                elif st_ == 'absent':
                    in_v = out_v = 'A'
                    fi = fo = ABS_FILL
                    absent_count += 1
                elif 'missing' in st_:
                    in_v  = dr['in_time']  or '?'
                    out_v = dr['out_time'] or '?'
                    fi = fo = MP_FILL
                else:
                    in_v  = dr['in_time']  or ''
                    out_v = dr['out_time'] or ''
                    fi    = IN_FILL
                    fo    = OUT_FILL
                    present_count += 1

                ot_v   = round(dr['ot_hrs'], 2) if dr['ot_hrs'] > 0 else ''
                ot_f   = WOT_FILL if is_wo and dr['ot_hrs'] > 0 else OT_FILL

                for r, val, fill in (
                    (row_num,   in_v,  fi),
                    (row_num+1, out_v, fo),
                    (row_num+2, ot_v,  ot_f),
                ):
                    cell = ws.cell(r, col, val if val != '' else None)
                    cell.fill      = fill
                    cell.alignment = center
                    cell.font      = Font(size=8)
                    cell.border    = border

                if dr['ot_type'] == 'regular':
                    reg_ot_total   += dr['ot_hrs']
                elif dr['ot_type'] == 'weekoff':
                    wkoff_ot_total += dr['ot_hrs']

            # Summary columns
            sc = 5 + num_days
            for r in (row_num, row_num+1, row_num+2):
                for c in range(sc, sc+4):
                    cell = ws.cell(r, c)
                    cell.border    = border
                    cell.alignment = center
                    cell.font      = Font(size=9)

            ws.cell(row_num, sc,   round(reg_ot_total,   2) or None)
            ws.cell(row_num, sc+1, round(wkoff_ot_total, 2) or None)
            ws.cell(row_num, sc+2, present_count or None)
            ws.cell(row_num, sc+3, absent_count  or None)

            for c in (sc, sc+1, sc+2, sc+3):
                ws.cell(row_num, c).font = Font(size=9, bold=True, color='1F4E79')

            row_num += 3
            # thin separator
            for col in range(1, sc + 4):
                ws.cell(row_num, col).fill = PatternFill('solid', fgColor='F5F5F5')
            ws.row_dimensions[row_num].height = 3
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_ot_summary_excel(emp_rows: list, year: int, month: int) -> bytes:
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = datetime(year, month, 1).strftime('%b-%Y OT')

    num_days  = calendar.monthrange(year, month)[1]
    dates     = [date(year, month, d) for d in range(1, num_days + 1)]

    HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=9)
    SUN_FILL  = PatternFill('solid', fgColor='EDE7F6')
    REG_FILL  = PatternFill('solid', fgColor='E8F5E9')
    WOT_FILL  = PatternFill('solid', fgColor='FFF9C4')
    TOT_FILL  = PatternFill('solid', fgColor='E3F2FD')
    border    = _thin_border()
    center    = _center()

    fixed_headers = ['Sl', 'Employee Name', 'Unit', 'Reg OT Rate\n(₹/hr)',
                     'Wkoff OT Rate\n(₹/hr)', 'Present', 'Absent']
    day_headers   = [str(d.day) + '\n' + d.strftime('%a') for d in dates]
    tail_headers  = ['Total Reg OT\n(hrs)', 'Total Wkoff OT\n(hrs)',
                     'Reg OT Amt\n(₹)', 'Wkoff OT Amt\n(₹)', 'TOTAL OT\n(₹)']
    all_headers   = fixed_headers + day_headers + tail_headers

    ws.row_dimensions[1].height = 32
    for col, h in enumerate(all_headers, 1):
        c = ws.cell(1, col, h)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = center
        c.border    = border
        if col >= 8 and col <= 7 + num_days:
            d = dates[col - 8]
            if d.weekday() == 6:
                c.fill = PatternFill('solid', fgColor='4A148C')

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    for col in range(8, 8 + num_days):
        ws.column_dimensions[get_column_letter(col)].width = 4.5
    for i in range(5):
        ws.column_dimensions[get_column_letter(8 + num_days + i)].width = 14

    ws.freeze_panes = 'H2'

    for idx, emp in enumerate(emp_rows, start=2):
        ot_rate   = emp['ot_rate'] or 0
        wkoff_rate = round(ot_rate * WO_MULT, 2)

        base = [idx - 1, emp['name'], emp['unit'], ot_rate, wkoff_rate,
                emp['present'], emp['absent']]
        for col, val in enumerate(base, 1):
            c = ws.cell(idx, col, val)
            c.alignment = center
            c.border    = border
            c.font      = Font(size=9)

        for d_idx, dr in enumerate(emp['day_records']):
            col   = 8 + d_idx
            val   = round(dr['ot_hrs'], 2) if dr['ot_hrs'] > 0 else None
            fill  = WOT_FILL if dr['is_weekoff'] and val else (REG_FILL if val else
                    (SUN_FILL if dr['status'] in ('sunday', 'holiday') else None))
            c     = ws.cell(idx, col, val)
            c.alignment = center
            c.border    = border
            c.font      = Font(size=8)
            if fill:
                c.fill = fill

        sc = 8 + num_days
        reg_pay   = round(emp['reg_ot']   * ot_rate,    2)
        wkoff_pay = round(emp['wkoff_ot'] * wkoff_rate, 2)
        total_pay = round(reg_pay + wkoff_pay,           2)

        for col, val in zip(range(sc, sc+5),
                            [emp['reg_ot'], emp['wkoff_ot'], reg_pay, wkoff_pay, total_pay]):
            c = ws.cell(idx, col, val if val else None)
            c.alignment = center
            c.border    = border
            c.font      = Font(size=9, bold=(col == sc + 4))
            if col == sc + 4:
                c.fill = TOT_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_misspunch_report(emp_rows: list, year: int, month: int) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Misspunch Report'

    HDR_FILL = PatternFill('solid', fgColor='C62828')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
    border   = _thin_border()
    center   = _center()

    headers = ['Sl', 'Employee Name', 'Unit', 'Date', 'Day',
               'Issue', 'Biometric IN', 'Biometric OUT',
               'Corrected IN', 'Corrected OUT', 'Action Taken']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = center; c.border = border

    widths = [4, 22, 8, 14, 8, 18, 12, 12, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    sl  = 1
    for emp in emp_rows:
        for dr in emp['misspunch_days']:
            issue = dr['status'].replace('_', ' ').title()
            for col, val in zip(range(1, 12), [
                sl, emp['name'], emp['unit'],
                dr['date'].strftime('%d-%b-%Y'), dr['day'],
                issue, dr['in_time'] or '—', dr['out_time'] or '—',
                '', '', 'Pending HR action'
            ]):
                c = ws.cell(row, col, val)
                c.alignment = center; c.border = border; c.font = Font(size=9)
            ws.cell(row, 9).fill = PatternFill('solid', fgColor='FFF9C4')
            ws.cell(row, 10).fill = PatternFill('solid', fgColor='FFF9C4')
            row += 1; sl += 1

    if row == 2:
        ws.cell(2, 1, '✅ No misspunches found for this month').font = Font(
            size=11, bold=True, color='2E7D32')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
#  APP HEADER
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div style="font-size:36px">🏭</div>
  <div>
    <h1>Excelhub HR Automation</h1>
    <p>Attendance · OT Calculation · Salary Processing &nbsp;|&nbsp; No coding required</p>
  </div>
</div>
""", unsafe_allow_html=True)

# Step bar
step_labels = ["① Upload Files", "② Review & Correct", "③ Download Reports"]
step_classes = []
for i in range(1, 4):
    if i < st.session_state.step:
        step_classes.append("step-item done")
    elif i == st.session_state.step:
        step_classes.append("step-item active")
    else:
        step_classes.append("step-item")

step_html = '<div class="step-bar">' + ''.join(
    f'<div class="{cls}">{lbl}</div>'
    for cls, lbl in zip(step_classes, step_labels)
) + '</div>'
st.markdown(step_html, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
#  STEP 1 — UPLOAD
# ─────────────────────────────────────────────────────────────
if st.session_state.step == 1:

    col_left, col_right = st.columns([1, 1], gap="large")

    with col_left:
        st.markdown("#### 📅 Select Month")
        m_col, y_col = st.columns(2)
        with m_col:
            month = st.selectbox("Month", list(range(1, 13)),
                                  index=st.session_state.month - 1,
                                  format_func=lambda x: datetime(2000, x, 1).strftime('%B'))
        with y_col:
            year = st.number_input("Year", 2020, 2035, value=st.session_state.year)

        st.session_state.month = month
        st.session_state.year  = year

        st.markdown("#### 🖥 Biometric .dat File")
        st.markdown('<div class="info-card"><h4>What is this?</h4>'
                    '<p>The raw log file exported from your biometric machine '
                    '(finger print / face scanner). Usually named like attendance.dat '
                    'or export_march.txt</p></div>', unsafe_allow_html=True)
        dat_file = st.file_uploader("Upload .dat / .txt file",
                                     type=['dat', 'txt', 'csv', 'log'],
                                     help="Raw biometric export file")
        if dat_file:
            content = dat_file.read().decode('utf-8', errors='ignore')
            st.session_state.parsed = parse_dat(content)
            n = len(st.session_state.parsed)
            if n:
                st.success(f"✅ Detected {n} employee IDs in the .dat file")
                with st.expander("Preview first 5 records"):
                    for i, (eid, dates_dict) in enumerate(
                            list(st.session_state.parsed.items())[:5]):
                        first_date = sorted(dates_dict.keys())[0]
                        times = [t.strftime('%H:%M')
                                 for t in dates_dict[first_date]]
                        st.code(f"ID: {eid}  |  {first_date}  →  {', '.join(times)}")
            else:
                st.warning("⚠️ Could not read any records. Check the file format.")

        st.markdown("#### 🗓 Public Holidays This Month")
        num_days  = calendar.monthrange(year, month)[1]
        all_dates = [date(year, month, d) for d in range(1, num_days + 1)]
        weekdays  = [d for d in all_dates if d.weekday() != 6]
        holidays  = st.multiselect(
            "Select government holidays (1.5× OT rate applies)",
            options=weekdays,
            format_func=lambda d: d.strftime('%d %b %Y — %A'),
            default=[]
        )
        st.session_state.holidays = holidays

    with col_right:
        st.markdown("#### 📊 Attendance Excel (Employee Name Mapping)")
        st.markdown('<div class="info-card"><h4>Why needed?</h4>'
                    '<p>Your biometric machine only stores numeric IDs. This Excel file '
                    '(Sheet3) maps those IDs to employee names and units.</p></div>',
                    unsafe_allow_html=True)
        att_file = st.file_uploader("Upload Attendance_Report Excel",
                                     type=['xlsx', 'xls'], key='att_up')
        if att_file:
            att_bytes = att_file.read()
            st.session_state.att_bytes = att_bytes
            emp_map = read_emp_map(att_bytes)
            st.session_state.emp_map = emp_map
            if emp_map:
                st.success(f"✅ Loaded {len(emp_map)} employees")
                with st.expander("Preview employee list"):
                    preview = [(v['name'], v['unit'], k)
                               for k, v in list(emp_map.items())[:10]]
                    st.dataframe(pd.DataFrame(preview,
                                 columns=['Name', 'Unit', 'Biometric ID']),
                                 hide_index=True, use_container_width=True)
            else:
                st.warning("No employees found in Sheet3.")

        st.markdown("#### 💰 Salary Excel (OT Rates)")
        st.markdown('<div class="info-card"><h4>Why needed?</h4>'
                    '<p>The Basic Details sheet has each employee\'s OT Per Hour rate. '
                    'This is used to calculate ₹ OT amounts automatically.</p></div>',
                    unsafe_allow_html=True)
        sal_file = st.file_uploader("Upload Excelhub_Salary_Computation Excel",
                                     type=['xlsx', 'xls'], key='sal_up')
        if sal_file:
            sal_bytes = sal_file.read()
            st.session_state.sal_bytes = sal_bytes
            if st.session_state.emp_map:
                st.session_state.emp_map = load_ot_rates(sal_bytes,
                                                          st.session_state.emp_map)
                mapped = sum(1 for v in st.session_state.emp_map.values()
                             if v.get('ot_rate', 0) > 0)
                st.success(f"✅ OT rates loaded for {mapped} employees")
            else:
                st.info("Upload Attendance Excel first to link OT rates.")

        st.markdown("#### 📋 OT Business Rules")
        st.markdown("""
<div class="rule-box">
  Total shift      = 12 hours<br>
  Lunch deduction  = 0.5 hrs (auto)<br>
  OT starts after  = 8 hrs net work<br>
  Max OT per day   = 3.3 hrs<br>
  Weekoff OT rate  = Regular × 1.5<br>
  Sunday / Holiday → Weekoff OT
</div>
""", unsafe_allow_html=True)

    st.divider()
    can_proceed = bool(st.session_state.emp_map)
    if not can_proceed:
        st.info("👆 Please upload at least the Attendance Excel to continue.")
    if st.button("Next → Review Attendance  ▶", type="primary",
                  disabled=not can_proceed, use_container_width=True):
        st.session_state.step = 2
        st.rerun()


# ─────────────────────────────────────────────────────────────
#  STEP 2 — REVIEW & CORRECT
# ─────────────────────────────────────────────────────────────
elif st.session_state.step == 2:

    month      = st.session_state.month
    year       = st.session_state.year
    month_name = datetime(year, month, 1).strftime('%B %Y')

    emp_rows = build_attendance(
        st.session_state.parsed,
        st.session_state.emp_map,
        year, month,
        st.session_state.holidays,
        st.session_state.corrections
    )

    # ── Metrics ──────────────────────────────────────────────
    total_emp    = len(emp_rows)
    total_mp     = sum(e['misspunch_count'] for e in emp_rows)
    total_pres   = sum(e['present']  for e in emp_rows)
    total_abs    = sum(e['absent']   for e in emp_rows)
    total_reg_ot = round(sum(e['reg_ot']   for e in emp_rows), 2)
    total_wo_ot  = round(sum(e['wkoff_ot'] for e in emp_rows), 2)

    st.markdown(f"### {month_name} — Attendance Review")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Employees",    total_emp)
    c2.metric("Present Days", total_pres)
    c3.metric("Absent Days",  total_abs)
    c4.metric("Regular OT",   f"{total_reg_ot} hrs")
    c5.metric("Weekoff OT",   f"{total_wo_ot} hrs")
    c6.metric("⚠️ Misspunches", total_mp)

    st.divider()

    # ── Misspunch corrections ─────────────────────────────────
    mp_emps = [e for e in emp_rows if e['misspunch_count'] > 0]
    if mp_emps:
        st.markdown("### ⚠️ Misspunch Corrections")
        st.caption("Fill actual IN/OUT times where biometric punch is missing. "
                   "Leave blank → system uses 8.5 hrs default (0 OT).")

        for emp in mp_emps:
            with st.expander(
                f"🔴  {emp['name']}  ·  {emp['unit']}  "
                f"·  {emp['misspunch_count']} misspunch day(s)"
            ):
                for dr in emp['misspunch_days']:
                    d_str   = str(dr['date'])
                    day_lbl = dr['date'].strftime('%d %b %Y  (%A)')
                    issue   = dr['status'].replace('_', ' ').upper()
                    st.markdown(f"**{day_lbl}** &nbsp; "
                                f'<span class="badge-miss">{issue}</span>',
                                unsafe_allow_html=True)

                    mc1, mc2, mc3 = st.columns([3, 3, 2])
                    corr_key = (emp['bio_id'], d_str)
                    existing = st.session_state.corrections.get(corr_key, (None, None))

                    with mc1:
                        in_v = st.text_input("Actual IN (HH:MM)",
                            value=existing[0].strftime('%H:%M')
                                  if existing[0] and hasattr(existing[0], 'strftime')
                                  else (existing[0] or ''),
                            key=f"in_{emp['bio_id']}_{d_str}",
                            placeholder="e.g. 09:05")
                    with mc2:
                        out_v = st.text_input("Actual OUT (HH:MM)",
                            value=existing[1].strftime('%H:%M')
                                  if existing[1] and hasattr(existing[1], 'strftime')
                                  else (existing[1] or ''),
                            key=f"out_{emp['bio_id']}_{d_str}",
                            placeholder="e.g. 21:15")
                    with mc3:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("✔ Save", key=f"save_{emp['bio_id']}_{d_str}"):
                            if in_v and out_v:
                                try:
                                    d   = dr['date']
                                    in_dt  = datetime.combine(
                                        d, datetime.strptime(in_v.strip(), '%H:%M').time())
                                    out_dt = datetime.combine(
                                        d, datetime.strptime(out_v.strip(), '%H:%M').time())
                                    st.session_state.corrections[corr_key] = (in_dt, out_dt)
                                    st.success("Saved!")
                                    st.rerun()
                                except Exception:
                                    st.error("Format must be HH:MM  e.g. 09:05")
                    st.markdown("---")
        st.divider()

    # ── Attendance summary table ──────────────────────────────
    st.markdown("### Attendance Summary")
    summary = [{
        'Name':          e['name'],
        'Unit':          e['unit'],
        'Present':       e['present'],
        'Absent':        e['absent'],
        'Reg OT (hrs)':  e['reg_ot'],
        'Wkoff OT (hrs)': e['wkoff_ot'],
        'OT Rate (₹)':   e['ot_rate'],
        'Reg OT (₹)':    round(e['reg_ot']   * e['ot_rate'],           2),
        'Wkoff OT (₹)':  round(e['wkoff_ot'] * e['ot_rate'] * WO_MULT, 2),
        'Misspunches':   e['misspunch_count'],
    } for e in emp_rows]

    st.dataframe(
        pd.DataFrame(summary),
        use_container_width=True,
        hide_index=True,
        column_config={
            'Reg OT (₹)':   st.column_config.NumberColumn(format="₹%.2f"),
            'Wkoff OT (₹)': st.column_config.NumberColumn(format="₹%.2f"),
            'OT Rate (₹)':  st.column_config.NumberColumn(format="₹%.0f"),
        }
    )

    st.divider()
    c_back, c_next = st.columns(2)
    with c_back:
        if st.button("◀ Back to Upload", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with c_next:
        if st.button("Next → Generate Reports  ▶", type="primary",
                      use_container_width=True):
            st.session_state.attendance = emp_rows
            st.session_state.step = 3
            st.rerun()


# ─────────────────────────────────────────────────────────────
#  STEP 3 — GENERATE & DOWNLOAD
# ─────────────────────────────────────────────────────────────
elif st.session_state.step == 3:

    emp_rows   = st.session_state.attendance or []
    month      = st.session_state.month
    year       = st.session_state.year
    month_name = datetime(year, month, 1).strftime('%B %Y')

    st.markdown(f"### ✅ {month_name} — Generate Reports")
    st.success(f"Ready to export reports for **{len(emp_rows)} employees**")

    # ── Anomaly summary ───────────────────────────────────────
    anomalies = []
    for e in emp_rows:
        if e['absent'] >= 15:
            anomalies.append(f"⚠️  {e['name']} ({e['unit']}) — {e['absent']} absent days")
        if e['misspunch_count'] > 0:
            anomalies.append(f"🔴  {e['name']} — {e['misspunch_count']} uncorrected misspunch(es)")
        if e['ot_rate'] == 0:
            anomalies.append(f"💡  {e['name']} — OT rate not found (check salary file)")

    if anomalies:
        with st.expander(f"⚠️  {len(anomalies)} items need attention — review before downloading"):
            for a in anomalies:
                st.markdown(f'<div class="anomaly-row">{a}</div>',
                            unsafe_allow_html=True)

    st.divider()

    # ── Download buttons ──────────────────────────────────────
    d1, d2, d3 = st.columns(3)

    with d1:
        st.markdown("#### 📋 Attendance Sheet")
        st.caption("IN · OUT · OT for all employees. Same format as your existing Excel.")
        if st.button("Generate Attendance Excel", use_container_width=True):
            with st.spinner("Building attendance sheet..."):
                data = generate_attendance_excel(emp_rows, year, month)
            st.download_button(
                "⬇️ Download Attendance Sheet",
                data=data,
                file_name=f"Attendance_{month_name.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with d2:
        st.markdown("#### 💰 OT Summary Sheet")
        st.caption("Regular OT + Weekoff OT amounts per employee. Paste into salary file.")
        if st.button("Generate OT Summary Excel", use_container_width=True):
            with st.spinner("Calculating OT..."):
                data = generate_ot_summary_excel(emp_rows, year, month)
            st.download_button(
                "⬇️ Download OT Summary",
                data=data,
                file_name=f"OT_Summary_{month_name.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with d3:
        st.markdown("#### 📌 Misspunch Report")
        st.caption("All flagged misspunches with space for HR to fill corrected times.")
        if st.button("Generate Misspunch Report", use_container_width=True):
            with st.spinner("Compiling report..."):
                data = generate_misspunch_report(emp_rows, year, month)
            st.download_button(
                "⬇️ Download Misspunch Report",
                data=data,
                file_name=f"Misspunch_{month_name.replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.divider()

    # ── Final stats ───────────────────────────────────────────
    st.markdown("#### Month Summary")
    s1, s2, s3, s4, s5 = st.columns(5)
    s1.metric("Total Employees", len(emp_rows))
    s2.metric("Total Present Days", sum(e['present']  for e in emp_rows))
    s3.metric("Total Absent Days",  sum(e['absent']   for e in emp_rows))
    s4.metric("Total Regular OT",
              f"{round(sum(e['reg_ot']   for e in emp_rows), 1)} hrs")
    s5.metric("Total Weekoff OT",
              f"{round(sum(e['wkoff_ot'] for e in emp_rows), 1)} hrs")

    st.divider()
    c_back, c_new = st.columns(2)
    with c_back:
        if st.button("◀ Back to Review", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with c_new:
        if st.button("🔄 Process New Month", use_container_width=True):
            for k, v in defaults.items():
                st.session_state[k] = v
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)
    st.caption("Built for Excelhub Technologies · Attendance & Payroll Automation")
